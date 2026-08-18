"""
Tests for services/excel_import.py - the row validation and file
parsing behind the "Import from Excel" feature.

Run with: pytest tests/test_excel_import.py
"""
import io
import sys
import os

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pytest
from openpyxl import Workbook

from services.excel_import import parse_import_file, validate_row, ImportFileError


# ---- validate_row --------------------------------------------------

def test_valid_row_has_no_errors():
    row = {"row_number": 2, "name": "Baby Wipes", "category": "Baby", "quantity": 19, "price": 17.6}
    result = validate_row(row)
    assert result["errors"] == {}
    assert result["name"] == "Baby Wipes"
    assert result["quantity"] == 19
    assert result["price"] == 17.6


def test_price_of_zero_is_valid():
    row = {"row_number": 2, "name": "Freebie", "category": "Misc", "quantity": 1, "price": 0}
    result = validate_row(row)
    assert result["errors"] == {}


def test_missing_name_is_flagged():
    row = {"row_number": 2, "name": "", "category": "Food", "quantity": 5, "price": 2}
    result = validate_row(row)
    assert "name" in result["errors"]


def test_missing_category_is_flagged():
    row = {"row_number": 2, "name": "Item", "category": None, "quantity": 5, "price": 2}
    result = validate_row(row)
    assert "category" in result["errors"]


def test_non_numeric_quantity_is_flagged():
    row = {"row_number": 2, "name": "Item", "category": "Food", "quantity": "abc", "price": 2}
    result = validate_row(row)
    assert "quantity" in result["errors"]


def test_negative_quantity_is_flagged():
    row = {"row_number": 2, "name": "Item", "category": "Food", "quantity": -5, "price": 2}
    result = validate_row(row)
    assert "quantity" in result["errors"]


def test_fractional_quantity_is_flagged():
    row = {"row_number": 2, "name": "Item", "category": "Food", "quantity": 5.5, "price": 2}
    result = validate_row(row)
    assert "quantity" in result["errors"]


def test_quantity_over_cap_is_flagged():
    row = {"row_number": 2, "name": "Item", "category": "Food", "quantity": 9999999999, "price": 2}
    result = validate_row(row)
    assert "quantity" in result["errors"]


def test_negative_price_is_flagged():
    row = {"row_number": 2, "name": "Item", "category": "Food", "quantity": 5, "price": -1}
    result = validate_row(row)
    assert "price" in result["errors"]


def test_non_numeric_price_is_flagged():
    row = {"row_number": 2, "name": "Item", "category": "Food", "quantity": 5, "price": "free"}
    result = validate_row(row)
    assert "price" in result["errors"]


def test_whitespace_padded_values_are_trimmed():
    row = {"row_number": 2, "name": "  Padded  ", "category": " Food ", "quantity": 5, "price": 2}
    result = validate_row(row)
    assert result["errors"] == {}
    assert result["name"] == "Padded"
    assert result["category"] == "Food"


def test_numeric_strings_are_accepted():
    row = {"row_number": 2, "name": "Item", "category": "Food", "quantity": "5", "price": "2.50"}
    result = validate_row(row)
    assert result["errors"] == {}
    assert result["quantity"] == 5
    assert result["price"] == 2.5


def test_completely_empty_quantity_is_flagged():
    row = {"row_number": 2, "name": "Item", "category": "Food", "quantity": None, "price": None}
    result = validate_row(row)
    assert "quantity" in result["errors"]


def test_invalid_row_returns_raw_value_for_display():
    """A bad cell should show the person what was actually there, not a
    coerced/default value, so they know what to fix."""
    row = {"row_number": 2, "name": "Item", "category": "Food", "quantity": "not-a-number", "price": 2}
    result = validate_row(row)
    assert result["quantity"] == "not-a-number"


# ---- parse_import_file ----------------------------------------------

class _FakeFileStorage:
    """Minimal stand-in for werkzeug's FileStorage - just needs .read()."""
    def __init__(self, data):
        self._data = data

    def read(self):
        return self._data


def _workbook_bytes(rows):
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def test_parses_valid_file_matching_expected_columns():
    data = _workbook_bytes([
        ["Name", "Category", "Quantity", "Price"],
        ["Baby Wipe", "baby", 19, 17.6],
        ["test3", "test3", 2, 6],
    ])
    rows = parse_import_file(_FakeFileStorage(data))
    assert len(rows) == 2
    assert rows[0]["name"] == "Baby Wipe"
    assert rows[0]["quantity"] == 19
    assert rows[0]["row_number"] == 2
    assert rows[1]["row_number"] == 3


def test_blank_trailing_rows_are_skipped():
    data = _workbook_bytes([
        ["Name", "Category", "Quantity", "Price"],
        ["Item", "Food", 5, 2],
        [None, None, None, None],
    ])
    rows = parse_import_file(_FakeFileStorage(data))
    assert len(rows) == 1


def test_header_matching_is_case_and_whitespace_insensitive():
    data = _workbook_bytes([
        [" name ", "CATEGORY", "quantity", "Price"],
        ["Widget", "Tools", 3, 9.99],
    ])
    rows = parse_import_file(_FakeFileStorage(data))
    assert len(rows) == 1
    assert rows[0]["name"] == "Widget"


def test_missing_required_columns_raises_import_file_error():
    data = _workbook_bytes([
        ["Name", "Category", "Qty"],  # "Qty" instead of "Quantity", no "Price"
        ["Item", "Food", 5],
    ])
    with pytest.raises(ImportFileError):
        parse_import_file(_FakeFileStorage(data))


def test_empty_file_raises_import_file_error():
    wb = Workbook()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    # A brand-new workbook has one empty sheet with no header row content
    # once the default cells are stripped - openpyxl still yields one row
    # from iter_rows for an empty sheet with no data, so this exercises
    # the "couldn't find headers" path either way.
    with pytest.raises(ImportFileError):
        parse_import_file(_FakeFileStorage(buf.read()))


def test_unreadable_file_raises_import_file_error():
    with pytest.raises(ImportFileError):
        parse_import_file(_FakeFileStorage(b"this is not a real xlsx file"))
