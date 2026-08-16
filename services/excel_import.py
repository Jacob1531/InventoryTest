import io

from openpyxl import load_workbook

from models import MAX_NUMERIC_VALUE

REQUIRED_COLUMNS = ["Name", "Category", "Quantity", "Price"]


class ImportFileError(Exception):
    """Raised when the uploaded file itself can't be read at all - wrong
    format, missing required columns, etc. Distinct from a row having bad
    values, which is handled per-row instead of failing the whole import."""
    pass


def parse_import_file(file_storage):
    """Reads an uploaded .xlsx file and returns a list of raw row dicts:
    [{"row_number": 2, "name": "...", "category": "...",
      "quantity": "...", "price": "..."}, ...]

    Values are returned as-is (not yet validated/coerced) so validate_row()
    can report exactly what was in the cell. Completely blank rows are
    skipped rather than treated as errors, since trailing blank rows are a
    common artifact of exporting/editing spreadsheets.
    """
    try:
        workbook = load_workbook(io.BytesIO(file_storage.read()), data_only=True, read_only=True)
    except Exception:
        raise ImportFileError(
            "Couldn't read that file. Make sure it's a valid .xlsx file."
        )

    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)

    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise ImportFileError("That file appears to be empty.")

    # Map lowercased/stripped header text -> column index
    header_map = {}
    for idx, cell in enumerate(header_row):
        if cell is not None:
            header_map[str(cell).strip().lower()] = idx

    missing = [col for col in REQUIRED_COLUMNS if col.lower() not in header_map]
    if missing:
        raise ImportFileError(
            "Missing required column(s): " + ", ".join(missing) +
            ". Expected headers: " + ", ".join(REQUIRED_COLUMNS) + "."
        )

    name_idx = header_map["name"]
    category_idx = header_map["category"]
    quantity_idx = header_map["quantity"]
    price_idx = header_map["price"]

    def cell_value(row, idx):
        if idx >= len(row):
            return None
        return row[idx]

    raw_rows = []
    for row_number, row in enumerate(rows_iter, start=2):
        name = cell_value(row, name_idx)
        category = cell_value(row, category_idx)
        quantity = cell_value(row, quantity_idx)
        price = cell_value(row, price_idx)

        # Skip fully blank rows (common trailing artifact in spreadsheets)
        if all(v is None or str(v).strip() == "" for v in (name, category, quantity, price)):
            continue

        raw_rows.append({
            "row_number": row_number,
            "name": name,
            "category": category,
            "quantity": quantity,
            "price": price,
        })

    return raw_rows


def _clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def _parse_number(value):
    """Returns a float if value is a valid number, else None. Accepts
    numbers already typed as int/float (openpyxl gives us these natively
    for numeric cells) as well as numeric-looking strings/text cells."""
    if value is None or (isinstance(value, str) and value.strip() == ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return None


def validate_row(raw_row):
    """Validates one raw row dict and returns:
    {"row_number", "name", "category", "quantity", "price", "errors"}

    For a field that fails validation, the corresponding output value is
    the original raw value (as a string) rather than a coerced one, so the
    person reviewing the import sees exactly what was in the cell.
    "errors" is a dict of field -> message; empty means the row is valid.
    """
    errors = {}

    name = _clean_text(raw_row.get("name"))
    if not name:
        errors["name"] = "Name is required."

    category = _clean_text(raw_row.get("category"))
    if not category:
        errors["category"] = "Category is required."

    quantity_num = _parse_number(raw_row.get("quantity"))
    if quantity_num is None:
        errors["quantity"] = "Quantity must be a whole number."
    elif quantity_num != int(quantity_num):
        errors["quantity"] = "Quantity must be a whole number."
    elif quantity_num < 0:
        errors["quantity"] = "Quantity can't be negative."
    elif quantity_num > MAX_NUMERIC_VALUE:
        errors["quantity"] = f"Quantity can't exceed {MAX_NUMERIC_VALUE}."

    price_num = _parse_number(raw_row.get("price"))
    if price_num is None:
        errors["price"] = "Price must be a number."
    elif price_num < 0:
        errors["price"] = "Price can't be negative."
    elif price_num > MAX_NUMERIC_VALUE:
        errors["price"] = f"Price can't exceed {MAX_NUMERIC_VALUE}."

    return {
        "row_number": raw_row.get("row_number"),
        "name": name if name else _clean_text(raw_row.get("name")),
        "category": category if category else _clean_text(raw_row.get("category")),
        "quantity": int(quantity_num) if quantity_num is not None and "quantity" not in errors else _clean_text(raw_row.get("quantity")),
        "price": price_num if price_num is not None and "price" not in errors else _clean_text(raw_row.get("price")),
        "errors": errors,
    }
